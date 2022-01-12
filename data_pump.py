import fdb
import openpyxl as opx
import time

import config

startTime = time.time()


def db_connect():
    con = fdb.connect(dsn=config.address, user=config.username, password=config.passwd,
                      charset='win1251')
    return con


c = db_connect()
cur = c.cursor()

inputData = []
wb = opx.load_workbook(filename=config.datasource)
sheet = wb.worksheets[0]

print('xlsx data read completed\n')

row_num = sheet.max_row
col_num = sheet.max_column

# TODO Re-write to use pandas df as an intermediate data storage?

data_l = [[None for t in range(10)] for k in range(row_num - 1)] # One row should be subtracted as data list
# does not have column headers. Otherwise one additional last empty row will be added to db

for r_num in range(2, row_num+1):
    data_l[r_num - 2][0] = str(sheet.cell(r_num, 1).value).strip()
    data_l[r_num - 2][1] = str(sheet.cell(r_num, 2).value).strip().encode('cp1251', errors='ignore')
    data_l[r_num - 2][2] = str(sheet.cell(r_num, 3).value).strip()
    data_l[r_num - 2][3] = str(sheet.cell(r_num, 4).value).strip()
    data_l[r_num - 2][4] = str(sheet.cell(r_num, 5).value).strip().encode('cp1251', errors='ignore')
    data_l[r_num - 2][5] = str(sheet.cell(r_num, 6).value).strip()
    """ date_in import """
    if str(sheet.cell(r_num, 7).value).strip() == 'None':
        data_l[r_num - 2][6] = None
    else:
        data_l[r_num - 2][6] = str(sheet.cell(r_num, 7).value).strip()
    data_l[r_num - 2][7] = str(sheet.cell(r_num, 8).value).strip().encode('cp1251', errors='ignore')
    """GBV import"""
    if sheet.cell(r_num, 9).value is None:
        data_l[r_num - 2][8] = None
    else:
        data_l[r_num - 2][8] = float(sheet.cell(r_num, 9).value)
    """NBV import"""
    if sheet.cell(r_num, 10).value is None:
        data_l[r_num - 2][9] = None
    else:
        data_l[r_num - 2][9] = float(sheet.cell(r_num, 10).value)

# with open(config.datasource, 'r', encoding='cp1251') as d_f:
#     for line in csv.reader(d_f):
#         print(line)
#         new_line = [s for s in line[0].split(';')]
#         print(new_line)
#         # if new_line[-1] == 'NBV':
#         #         #     continue
#         try:
#             new_line[-1] = float(new_line[-1])
#         except ValueError:
#             new_line[-1] = 0.0
#         try:
#             new_line[-2] = float(new_line[-2])
#         except ValueError:
#             new_line[-2] = 0.0
#         print(new_line)
#         new_line[6] = datetime.datetime.date(new_line[6])
#         inputData.append(new_line)

#
insert_statement = cur.prep(config.data_pump_req)

counts = 0
for row in data_l:
    # print(row)
    cur.execute(insert_statement, row)
    counts += 1
    print(counts)
    # if counts % 1000 == 0:
    #     c.commit()
    #     print(f'{counts} records committed')

c.commit()
c.close()
print(f'data pump completed, execution time: {time.time() - startTime} seconds')

# r = cur.fetchall()
# for s in r:
#     print(s)
