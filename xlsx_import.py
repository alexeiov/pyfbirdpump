import openpyxl as opx
import config
from file_path_getter import get_path


class XlsxImport:
    pass


# open_path = input('Enter path to file:')
# """File should be in xlsx format"""
# file_to_process = input('Enter filename:')
# full_data_path = Path(open_path).joinpath(file_to_process)

full_data_path = get_path()


def get_xlsx_data(data_path=full_data_path, data_workbook_num=config.wb_num):
    wb = opx.load_workbook(data_path)
    sheet = wb.worksheets[data_workbook_num]
    print('xlsx data read completed\n')
    return sheet


def make_data_dict(xlsx_sheet, start_col_num=1, nums_number=10):
    data_d = {}
    row_num = xlsx_sheet.max_row
    # col_num = xlsx_sheet.max_column
    for r_num in range(2, row_num+1):
        # print(str(xlsx_sheet.cell(r_num, col_num).value))
        if xlsx_sheet.cell(r_num, start_col_num).value is not None:
            data_d[str(xlsx_sheet.cell(r_num, start_col_num).value)] = str(xlsx_sheet.cell(r_num, start_col_num + 1).value)
    return data_d


# TODO Make data dict creation function for arbitrary number of columns
def make_data_list_mult_col(xlsx_sheet, start_col_num=1):
    data_d = []
    row_num = xlsx_sheet.max_row
    # col_num = xlsx_sheet.max_column
    for r_num in range(2, row_num+1):
        # print(str(xlsx_sheet.cell(r_num, col_num).value))
        if xlsx_sheet.cell(r_num, start_col_num).value is not None:
            data_d.append((str(xlsx_sheet.cell(r_num, start_col_num).value), str(xlsx_sheet.cell(r_num, start_col_num + 1).value),
                                                                        str(xlsx_sheet.cell(r_num, start_col_num + 2).value),
                           str(xlsx_sheet.cell(r_num, start_col_num + 3).value),
                           str(xlsx_sheet.cell(r_num, start_col_num + 4).value),
                           )) # add more columns to tuple to extend dataset. ef xlsx file contains less columins than here, empty columns will be filled with None, function will still work
    return data_d


if __name__ == "__main__":
    ws = get_xlsx_data(full_data_path)
    data = make_data_list_mult_col(ws)
