import openpyxl as opx
import config
from pathlib import Path
from connection import db_connect
import datetime
import re

db_name = config.address[-17:-4]
# base_filename = '5856_VSMPO_calculation_example_'
exported_filename = db_name + '_calculation_example_' + datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S') + '.xlsx' #Todo Get DB name from db
save_path = input('Please insert save path: ')
full_save_path = Path(save_path).joinpath(exported_filename)


def get_names(req):
    # procedure_field_names = input('Insert request to db to get field names: ')
    procedure_field_names = req
    sheet_name = get_table_name(req)
    proc_names = []
    c = db_connect()
    cur = c.cursor()
    cur.execute(procedure_field_names)
    names = cur.fetchall()
    for name in names:
        if name[0][0:2] != 'IP':
            proc_names.append(name[0].strip())
    return proc_names


def get_data(req):
    # data_get_example = input('Insert request to db to get valuation data: ')
    data_get_example = req
    data = []
    sheet_name = get_table_name(req)
    c = db_connect()
    cur = c.cursor()
    cur.execute(data_get_example)
    print('Export is in progress')
    # k = 0
    for line in cur.fetchall():
        # print(line)
        # while k < 200:
        #     print('.', end='')
        data.append(line)
        # k += 1
    # print('\n')
    c.close()
    return sheet_name, data


# def get_coeffs(req):
#     data = []
#     c = db_connect()
#     cur = c.cursor()
#     cur.execute(req)
#     for line in cur.fetchall():
#         data.append(line)
#     return data

def get_table_name(req):
    name_pattern = r'from (\S+)'
    table_name = re.search(name_pattern, req).group(1)
    return table_name


def save_results_to_xlsx(save_p, *res_sets): # Todo accept any number of sheets, first with formulas, other - simple export

    if not len(res_sets) % 2:
        pass
    else:
        print('Number of input result sets should be even, but it is not. Please correct')
        return

    P_CRN_PREV = 'G'
    P_TREND_COEFF = 'N'
    P_IDC_COEFF_PREV = 'J'
    P_DIRECT_COEFF = 'K'
    P_INDIRECT_COEFF = 'L'
    P_DIRECT_COEFF_PREV = 'H'
    P_INDIRECT_COEFF_PREV = 'I'
    P_IDC_COEFF = 'M'
    P_EURO_PREV = 'Q'
    P_EURO_CURRENT = 'S'
    P_USD_PREV = 'P'
    P_USD_CURRENT = 'R'

    CRN_CALC_C_NUM = 21

    wb = opx.Workbook()
    # wb.worksheets[0].title = 'calculation_example'

    for sheet_num in range(int(len(res_sets)/2)):
        wb.create_sheet(str(sheet_num))

    for sheet_num, res_set in enumerate(res_sets[int(len(res_sets)/2):]):
        sheet_name = res_set[0]
        wb.worksheets[sheet_num].title = str(sheet_name)

    # sheet_main_title = 'calculation_example'
    # wb.create_sheet(sheet_main_title, 0)
    # sheet_coeffs_prev_title = 'crn_trend_coeffs'
    # wb.create_sheet(sheet_coeffs_prev_title, 1)

    # names_r = res_sets[0]
    # names_c = res_sets[1]
    # names_t = res_sets[2]
    # results = res_sets[3][1]
    # coeffs = res_sets[4][1]
    # trends = res_sets[5][1]

    for num, data in enumerate(zip(res_sets[:int(len(res_sets)/2)],res_sets[int(len(res_sets)/2):])):

        for n_num, name in enumerate(data[0]):
            wb.worksheets[num].cell(1, n_num + 1).value = name

        for r_num, row in enumerate(data[1][1]):
            if not num:
                for cell_num, cell in enumerate(row):
                    wb.worksheets[num].cell(r_num + 2, cell_num + 1).value = cell
                    if row[2] in ('0', '1', '2', '3') and row[14][:10] == 'index_prev':
                        wb.worksheets[num].cell(r_num + 2,
                                              CRN_CALC_C_NUM).value = f'= {P_CRN_PREV}{r_num + 2} / {P_IDC_COEFF_PREV}{r_num + 2} * {P_IDC_COEFF}{r_num + 2} * {P_TREND_COEFF}{r_num + 2}'
                    elif row[2] not in ('0', '1', '2', '3', 'CIP') and row[14][:2] == 'ru':
                        wb.worksheets[num].cell(r_num + 2,
                                              CRN_CALC_C_NUM).value = f'= {P_CRN_PREV}{r_num + 2} / {P_IDC_COEFF_PREV}{r_num + 2} / {P_DIRECT_COEFF_PREV}{r_num + 2} / {P_INDIRECT_COEFF_PREV}{r_num + 2} * {P_IDC_COEFF}{r_num + 2} * {P_TREND_COEFF}{r_num + 2} * {P_INDIRECT_COEFF}{r_num + 2} * {P_DIRECT_COEFF}{r_num + 2}'
                    elif row[2] not in ('0', '1', '2', '3', 'CIP') and (row[14][:2] == 'eu' or row[14] == 'auto_imp') and \
                            row[1][:5] != 'Tirus':
                        pass
                    elif row[2] not in ('0', '1', '2', '3', 'CIP') and row[14] in ('us_tool', 'jap_tool', 'china') and row[
                                                                                                                           1][
                                                                                                                       :5] != 'Tirus':
                        pass
            else:
                for cell_num, cell in enumerate(row):
                    wb.worksheets[num].cell(r_num + 2, cell_num + 1).value = cell

    # for n_num, name in enumerate(names_r):
    #     wb.worksheets[0].cell(1, n_num + 1).value = name
    #
    # for r_num, row in enumerate(results):
    #     for cell_num, cell in enumerate(row):
    #         wb.worksheets[0].cell(r_num + 2, cell_num + 1).value = cell
    #         if row[2] in ('0', '1', '2', '3') and row[14][:10] == 'index_prev':
    #             wb.worksheets[0].cell(r_num + 2, CRN_CALC_C_NUM).value = f'= {P_CRN_PREV}{r_num + 2} / {P_IDC_COEFF_PREV}{r_num + 2} * {P_IDC_COEFF}{r_num + 2} * {P_TREND_COEFF}{r_num + 2}'
    #         elif row[2] not in ('0', '1', '2', '3', 'CIP') and row[14][:2] == 'ru':
    #             wb.worksheets[0].cell(r_num + 2, CRN_CALC_C_NUM).value = f'= {P_CRN_PREV}{r_num + 2} / {P_IDC_COEFF_PREV}{r_num + 2} / {P_DIRECT_COEFF_PREV}{r_num + 2} / {P_INDIRECT_COEFF_PREV}{r_num + 2} * {P_IDC_COEFF}{r_num + 2} * {P_TREND_COEFF}{r_num + 2} * {P_INDIRECT_COEFF}{r_num + 2} * {P_DIRECT_COEFF}{r_num + 2}'
    #         elif row[2] not in ('0', '1', '2', '3', 'CIP') and (row[14][:2] == 'eu' or row[14] == 'auto_imp') and row[1][:5] != 'Tirus':
    #             pass
    #         elif row[2] not in ('0', '1', '2', '3', 'CIP') and row[14] in ('us_tool', 'jap_tool', 'china') and row[1][:5] != 'Tirus':
    #             pass
    #
    # for n_num, name in enumerate(names_c):
    #     wb.worksheets[1].cell(1, n_num + 1).value = name
    #
    # for r_num, row in enumerate(coeffs):
    #     for cell_num, cell in enumerate(row):
    #         wb.worksheets[1].cell(r_num + 2, cell_num + 1).value = cell
    #
    # for n_num, name in enumerate(names_t):
    #     wb.worksheets[2].cell(1, n_num + 1).value = name
    #
    # for r_num, row in enumerate(trends):
    #     for cell_num, cell in enumerate(row):
    #         wb.worksheets[2].cell(r_num + 2, cell_num + 1).value = cell

    wb.save(filename=save_p)
    print('Export completed')


def sv_round(num):
    pass


# print(get_table_name("select * from calc_example('VSMPO') rows 7000"))

n_d = get_names("select RDB$parameter_name from rdb$procedure_parameters where rdb$procedure_name = 'CALC_EXAMPLE' order by rdb$parameter_number")
n_c = get_names("select rdb$field_name from rdb$relation_fields where rdb$relation_name = 'TREND_COEFFS' order by RDB$FIELD_POSITION")
n_t = get_names("select rdb$field_name from rdb$relation_fields where rdb$relation_name = 'COEFFS_GBV' order by RDB$FIELD_POSITION")
d = get_data("select * from calc_example('VSMPO') rows 7000")
c = get_data("select * from TREND_COEFFS")
e = get_data("select * from COEFFS_GBV")
save_results_to_xlsx(full_save_path, *(n_d, n_c, n_t, d, c, e,))
